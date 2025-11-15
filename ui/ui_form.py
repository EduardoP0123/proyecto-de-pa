import os
import sys
import threading
import tkinter as tk
from pathlib import Path
import math
import re
from datetime import datetime, timedelta
from tkinter import ttk, filedialog, messagebox
from tkcalendar import DateEntry
from collections import defaultdict
import pandas as pd

try:
    from PIL import Image, ImageTk  
except Exception:
    Image = ImageTk = None

project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root))
try:
    from src.ui_components import run_ui
except Exception:
    run_ui = None
from src.csv_processor import CSVProcessor


MONTH_ABBR_ES = {
    1: "ene", 2: "feb", 3: "mar", 4: "abr", 5: "may", 6: "jun",
    7: "jul", 8: "ago", 9: "sep", 10: "oct", 11: "nov", 12: "dic"
}

def format_es_date(d: datetime.date) -> str:
    # 1-ago-25
    return f"{d.day}-{MONTH_ABBR_ES.get(d.month, '')}-{d.strftime('%y')}"

class CSVUploaderApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Lecturas KV2C - v2.0")
        # Escalado para pantallas FHD/4K
        try:
            self.root.tk.call("tk", "scaling", 1.25)
        except Exception:
            pass

        self._init_style()

        self.root.geometry("980x680")
        self.root.minsize(900, 600)
        self.root.rowconfigure(0, weight=1)
        self.root.columnconfigure(0, weight=1)

        # referencia del logo para evitar GC
        self.seg_logo_img = None

        # Procesador
        workspace_path = Path.home() / "Downloads" / "BILLREAD_WORKSPACE"
        workspace_path.mkdir(parents=True, exist_ok=True)
        self.csv_processor = run_ui(workspace_path) if run_ui else CSVProcessor(workspace_path)

        # UI
        self.create_widgets()
        self._build_statusbar()

        self.company_multipliers = {}  # cache por empresa
        self.last_report = None        # (df, totals, meta)

    # ---------- Estilo ----------
    def _init_style(self):
        style = ttk.Style()
        # Tema estable y moderno
        try:
            style.theme_use("clam")
        except Exception:
            pass
        # Colores base
        self.COLOR_BG = "#0D1B2A"   # azul oscuro header
        self.COLOR_ACCENT = "#1B6F9B"
        self.COLOR_BTN = "#1F7A8C"

        style.configure("Header.TFrame", background=self.COLOR_BG)
        style.configure("Header.TLabel", background=self.COLOR_BG, foreground="white", font=("Segoe UI", 16, "bold"))
        style.configure("TLabel", font=("Segoe UI", 10))
        style.configure("TButton", font=("Segoe UI", 10))
        style.configure("Accent.TButton", font=("Segoe UI", 10, "bold"), foreground="white", background=self.COLOR_BTN)
        style.map("Accent.TButton", background=[("active", "#2390A6"), ("disabled", "#9fb6bf")])
        style.configure("Card.TFrame", padding=12)
        style.configure("Section.TLabelframe", padding=12)
        style.configure("Section.TLabelframe.Label", font=("Segoe UI", 11, "bold"))

    # ---------- Logo ----------
    def _load_seg_logo(self, max_h=56, max_w=220):
        """Carga y escala el logo buscando en ui/images y assets."""
        try:
            here = Path(__file__).resolve().parent
            roots = [
                here / "images",                    
                here.parent / "assets",             
                Path.cwd() / "ui" / "images",       
                Path.cwd() / "assets",
            ]
            candidates = [
                "seg.png", "SEG.png", "seg_logo.png",
                "seg.jpg", "SEG.jpg", "seg_logo.jpg",
                "seg.gif", "SEG.gif"
            ]
            for root in roots:
                for name in candidates:
                    p = root / name
                    if not p.exists():
                        continue
                    # Con Pillow (admite JPG/PNG/GIF y mejor escalado)
                    if Image and ImageTk:
                        img = Image.open(p).convert("RGBA")
                        r = min(max_h / img.height, max_w / img.width, 1.0)
                        new_size = (max(1, int(img.width * r)), max(1, int(img.height * r)))
                        img = img.resize(new_size, Image.LANCZOS)
                        return ImageTk.PhotoImage(img)
                    # Fallback sin Pillow: PhotoImage (PNG/GIF)
                    pic = tk.PhotoImage(file=str(p))
                    h, w = pic.height(), pic.width()
                    factor = max(1, math.ceil(max(h / max_h, w / max_w)))
                    if factor > 1:
                        pic = pic.subsample(factor, factor)
                    return pic
        except Exception as e:
            try:
                self.append_info(f"[LOGO] No se pudo cargar: {e}")
            except Exception:
                pass
        # Aviso si no se encontró
        try:
            self.append_info("[LOGO] No se encontró imagen en ui/images o assets (ej: ui/images/seg.png).")
        except Exception:
            pass
        return None

    # ---------- Layout ----------
    def create_widgets(self):
        root_frame = ttk.Frame(self.root, padding=0)
        root_frame.grid(row=0, column=0, sticky="nsew")
        root_frame.rowconfigure(1, weight=1)
        root_frame.columnconfigure(0, weight=1)

        # Header
        header = ttk.Frame(root_frame, style="Header.TFrame", padding=(16, 12))
        header.grid(row=0, column=0, sticky="ew")
        header.columnconfigure(0, weight=1)

        ttk.Label(header, text="Lecturas KV2C / KV2A analyzer", style="Header.TLabel").grid(row=0, column=0, sticky="w")
        self.seg_logo_img = self._load_seg_logo()
        if self.seg_logo_img:
            ttk.Label(header, image=self.seg_logo_img, style="Header.TLabel").grid(row=0, column=1, sticky="e")
        else:
            ttk.Label(header, text="SEG", style="Header.TLabel").grid(row=0, column=1, sticky="e")

        # Body: panel izquierdo opciones, derecho log
        body = ttk.Frame(root_frame, padding=12)
        body.grid(row=1, column=0, sticky="nsew")
        body.columnconfigure(0, weight=0)
        body.columnconfigure(1, weight=1)
        body.rowconfigure(0, weight=1)

        # Panel opciones (card)
        opts = ttk.Labelframe(body, text="Opciones", style="Section.TLabelframe")
        opts.grid(row=0, column=0, sticky="ns", padx=(0, 12))
        for i in range(6):
            opts.columnconfigure(i, weight=0)

        # Carpeta
        ttk.Label(opts, text="Carpeta").grid(row=0, column=0, sticky="w")
        self.folder_path = ttk.Entry(opts, width=40)
        self.folder_path.grid(row=0, column=1, columnspan=4, sticky="ew", padx=8)
        ttk.Button(opts, text="Examinar...", command=self.browse_folder).grid(row=0, column=5, sticky="e")

        # Resolución
        ttk.Label(opts, text="Resolución").grid(row=1, column=0, sticky="w", pady=(8, 0))
        self.resolution = ttk.Combobox(opts, values=["15min", "1h"], state="readonly", width=8)
        self.resolution.set("15min")
        self.resolution.grid(row=1, column=1, sticky="w", padx=8, pady=(8, 0))

        # Rango fechas/horas
        ttk.Label(opts, text="Inicio").grid(row=2, column=0, sticky="w", pady=(8, 0))
        self.start_date = DateEntry(opts, date_pattern="dd/mm/y", width=10)
        self.start_date.grid(row=2, column=1, sticky="w", padx=(8, 4), pady=(8, 0))
        self.start_hour = ttk.Spinbox(opts, from_=0, to=23, width=3)
        self.start_hour.set("00")
        self.start_hour.grid(row=2, column=2, sticky="w", pady=(8, 0))
        self.start_min = ttk.Spinbox(opts, from_=0, to=59, width=3)
        self.start_min.set("00")
        self.start_min.grid(row=2, column=3, sticky="w", padx=(4, 0), pady=(8, 0))

        ttk.Label(opts, text="Fin").grid(row=3, column=0, sticky="w")
        self.end_date = DateEntry(opts, date_pattern="dd/mm/y", width=10)
        self.end_date.grid(row=3, column=1, sticky="w", padx=(8, 4))
        self.end_hour = ttk.Spinbox(opts, from_=0, to=23, width=3)
        self.end_hour.set("23")
        self.end_hour.grid(row=3, column=2, sticky="w")
        self.end_min = ttk.Spinbox(opts, from_=0, to=59, width=3)
        self.end_min.set("59")
        self.end_min.grid(row=3, column=3, sticky="w", padx=(4, 0))

        # Empresa y Multiplo
        ttk.Label(opts, text="Empresa").grid(row=4, column=0, sticky="w", pady=(8, 0))
        self.company_cb = ttk.Combobox(opts, values=[], state="disabled", width=28)
        self.company_cb.grid(row=4, column=1, columnspan=3, sticky="ew", padx=(8, 0), pady=(8, 0))
        # Cuando cambia la empresa, sincronizar el multiplo mostrado
        self.company_cb.bind('<<ComboboxSelected>>', self._on_company_selected)

        ttk.Label(opts, text="Multiplo").grid(row=5, column=0, sticky="w")
        self.multiplier_sp = ttk.Spinbox(opts, from_=1, to=100000, width=8)
        self.multiplier_sp.set("80")
        self.multiplier_sp.grid(row=5, column=1, sticky="w", padx=(8, 0))
        # Guardar y propagar cambios de multiplo
        try:
            self.default_multiplier = 80
        except Exception:
            self.default_multiplier = 80
        self.multiplier_sp.configure(command=self._on_multiplier_change)
        self.multiplier_sp.bind('<FocusOut>', lambda e: self._on_multiplier_change())
        self.multiplier_sp.bind('<Return>', lambda e: self._on_multiplier_change())

        # Botonera
        btns = ttk.Frame(opts)
        btns.grid(row=6, column=0, columnspan=6, sticky="ew", pady=(12, 0))
        btns.columnconfigure(0, weight=1)
        ttk.Button(btns, text="Analizar CSV", style="Accent.TButton", command=self.analyze_folder).grid(row=0, column=0, sticky="ew")
        ttk.Button(btns, text="Analizar PRN", command=self.analyze_folder_prn).grid(row=0, column=1, sticky="ew", padx=(8, 0))
        # Exportaciones
        self.export_excel_btn = ttk.Button(btns, text="Exportar Excel", command=self.export_excel, state="disabled")
        self.export_excel_btn.grid(row=1, column=0, sticky="ew", pady=(8, 0))
        self.export_csv_btn = ttk.Button(btns, text="Exportar CSV", command=self.export_csv, state="disabled")
        self.export_csv_btn.grid(row=1, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))
        # Limpiar y reporte
        self.clear_btn = ttk.Button(btns, text="Limpiar", command=self.clear_results)
        self.clear_btn.grid(row=2, column=0, sticky="ew", pady=(8, 0))
        self.report_btn = ttk.Button(btns, text="Generar reporte mensual", command=self.generate_report, state="disabled")
        self.report_btn.grid(row=2, column=1, sticky="ew", padx=(8, 0), pady=(8, 0))

        # Panel de resultados (log)
        right = ttk.Labelframe(body, text="Registro y resultados", style="Section.TLabelframe")
        right.grid(row=0, column=1, sticky="nsew")
        right.rowconfigure(0, weight=1)
        right.columnconfigure(0, weight=1)

        self.info_text = tk.Text(right, height=20, wrap="word", font=("Consolas", 10))
        yscroll = ttk.Scrollbar(right, orient="vertical", command=self.info_text.yview)
        self.info_text.configure(yscrollcommand=yscroll.set)
        self.info_text.grid(row=0, column=0, sticky="nsew")
        yscroll.grid(row=0, column=1, sticky="ns")

        # Permitir expandir el panel derecho
        body.rowconfigure(0, weight=1)
        body.columnconfigure(1, weight=1)

    def _build_statusbar(self):
        bar = ttk.Frame(self.root, padding=(12, 6))
        bar.grid(row=2, column=0, sticky="ew")
        bar.columnconfigure(0, weight=1)
        self.status_label = ttk.Label(bar, text="Listo")
        self.status_label.grid(row=0, column=0, sticky="w")
        self.progress = ttk.Progressbar(bar, mode="indeterminate", length=160)
        self.progress.grid(row=0, column=1, sticky="e")

    # ---------- Utilidades UI ----------
    def set_busy(self, busy: bool, msg: str = ""):
        if busy:
            self.status_label.config(text=msg or "Procesando…")
            try:
                self.progress.start(12)
            except Exception:
                pass
        else:
            try:
                self.progress.stop()
            except Exception:
                pass
            self.status_label.config(text=msg or "Listo")

    def append_info(self, text):
        self.info_text.configure(state="normal")
        self.info_text.insert("end", text + "\n")
        self.info_text.see("end")
        self.info_text.configure(state="disabled")

    def show_error(self, message: str):
        try:
            messagebox.showerror("Error", message)
        finally:
            try:
                self.append_info(f"[ERROR] {message}")
            except Exception:
                pass

    def browse_folder(self):
        folder = filedialog.askdirectory(title="Selecciona carpeta con archivos")
        if folder:
            self.folder_path.delete(0, tk.END)
            self.folder_path.insert(0, folder)

    # ---------- Saneo de horas/minutos ----------
    def _sanitize_time_inputs(self):
        def to_int(v, default):
            try:
                return int(str(v).strip())
            except Exception:
                return default
        def clamp(v, lo, hi):
            return max(lo, min(hi, v))
        sh = clamp(to_int(self.start_hour.get(), 0), 0, 23)
        sm = clamp(to_int(self.start_min.get(), 0), 0, 59)
        eh = clamp(to_int(self.end_hour.get(), 23), 0, 23)
        em = clamp(to_int(self.end_min.get(), 59), 0, 59)
        # Reflejar en la UI por si el usuario puso 60, etc.
        try:
            self.start_hour.set(f"{sh:02d}")
            self.start_min.set(f"{sm:02d}")
            self.end_hour.set(f"{eh:02d}")
            self.end_min.set(f"{em:02d}")
        except Exception:
            pass
        return sh, sm, eh, em

    # ---------- Acciones ----------
    def analyze_folder(self):
        self._run_analysis("csv")

    def analyze_folder_prn(self):
        self._run_analysis("prn")

    def _run_analysis(self, file_type="csv"):
        folder_path = self.folder_path.get()
        if not folder_path or not os.path.exists(folder_path):
            messagebox.showerror("Error", "Selecciona una carpeta válida")
            return

        # Fechas y horas
        sdate = self.start_date.get_date()
        edate = self.end_date.get_date()
        sh, sm, eh, em = self._sanitize_time_inputs()
        start_dt = datetime(sdate.year, sdate.month, sdate.day, sh, sm)
        end_dt = datetime(edate.year, edate.month, edate.day, eh, em)
        if end_dt < start_dt:
            messagebox.showerror("Error", "El fin debe ser posterior al inicio")
            return

        resolution = self.resolution.get()

        # Preparar UI
        self.info_text.configure(state="normal")
        self.info_text.delete("1.0", "end")
        self.info_text.configure(state="disabled")
        self.set_busy(True, "Procesando…")

        def progress_cb(msg: str):
            self.root.after(0, lambda: self.append_info(msg))

        def month_span(s, e):
            y, m = s.year, s.month
            while (y < e.year) or (y == e.year and m <= e.month):
                yield y, m
                m += 1
                if m > 12:
                    m = 1
                    y += 1

        def worker():
            try:
                months = list(month_span(sdate, edate))
                monthly_dfs, all_details = [], []
                total_files = total_processed = total_errors = 0
                last_folder = folder_path

                for (yy, mm) in months:
                    # Límites de hora por mes iterado
                    if (yy, mm) == (sdate.year, sdate.month) and (yy, mm) == (edate.year, edate.month):
                        s_t = f"{sh:02d}:{sm:02d}"
                        e_t = f"{eh:02d}:{em:02d}"
                    elif (yy, mm) == (sdate.year, sdate.month):
                        s_t, e_t = f"{sh:02d}:{sm:02d}", "23:59"
                    elif (yy, mm) == (edate.year, edate.month):
                        s_t, e_t = "00:00", f"{eh:02d}:{em:02d}"
                    else:
                        s_t, e_t = "00:00", "23:59"

                    if file_type == "csv":
                        ok, msg, results = self.csv_processor.analyze_folder(
                            Path(folder_path), mes_usuario=mm, año_usuario=yy,
                            start_time=s_t, end_time=e_t, progress_cb=progress_cb
                        )
                    else:
                        if hasattr(self.csv_processor, "analyze_folder_prn"):
                            ok, msg, results = self.csv_processor.analyze_folder_prn(
                                Path(folder_path), mes_usuario=mm, año_usuario=yy,
                                start_time=s_t, end_time=e_t, progress_cb=progress_cb
                            )
                        else:
                            ok, msg, results = False, "Función PRN no disponible", None

                    if ok and getattr(self.csv_processor, "combined_df", None) is not None:
                        monthly_dfs.append(self.csv_processor.combined_df.copy())
                        all_details.extend(results.get("file_details", []))
                        total_files += results.get("total_files", 0)
                        total_processed += results.get("processed_files", 0)
                        total_errors += results.get("error_files", 0)
                        last_folder = results.get("folder", last_folder)

                if not monthly_dfs:
                    self.root.after(0, lambda: self.set_busy(False, "Sin datos"))
                    self.root.after(0, lambda: self.append_info("No se generaron datos"))
                    return

                combined = pd.concat(monthly_dfs, ignore_index=True)

                # Normalizar timestamp si no es datetime
                if "timestamp" in combined.columns and not pd.api.types.is_datetime64_any_dtype(combined["timestamp"]):
                    # Intentos múltiples de parseo
                    combined["timestamp"] = pd.to_datetime(
                        combined["timestamp"].astype(str).str.strip(),
                        errors="coerce", dayfirst=True
                    )
                # Eliminar filas sin timestamp válido
                if "timestamp" in combined.columns:
                    before_rows = combined.shape[0]
                    combined = combined[pd.notna(combined["timestamp"])].copy()
                    after_parse_rows = combined.shape[0]
                else:
                    before_rows = combined.shape[0]
                    after_parse_rows = before_rows

                # Filtro por rango final
                if "timestamp" in combined.columns:
                    combined = combined[(combined["timestamp"] >= start_dt) & (combined["timestamp"] <= end_dt)]
                after_filter_rows = combined.shape[0]

                if resolution == "1h" and not combined.empty and "timestamp" in combined.columns:
                    combined["hour_ts"] = combined["timestamp"].dt.floor("H")
                    agg_cols = {c: "sum" for c in ["kwh", "kvarh"] if c in combined.columns}
                    grouped = combined.groupby(["company", "hour_ts"], as_index=False).agg(agg_cols)
                    combined = grouped.rename(columns={"hour_ts": "timestamp"})
                    combined = combined.sort_values(["company", "timestamp"]).reset_index(drop=True)

                self.csv_processor.combined_df = combined

                # Agregados y detalles
                fmt = "%d/%m/%y %H:%M"
                agg = defaultdict(lambda: {
                    "filename": None, "rows": 0, "success": False,
                    "kwh_values": 0, "kvar_values": 0,
                    "start_date": None, "end_date": None, "error": None
                })
                for d in all_details:
                    name = d.get("filename")
                    if not name:
                        continue
                    a = agg[name]
                    a["filename"] = name
                    a["rows"] += int(d.get("rows", 0))
                    a["success"] = a["success"] or bool(d.get("success", False))
                    a["kwh_values"] += int(d.get("kwh_values", 0))
                    a["kvar_values"] += int(d.get("kvar_values", 0))
                dedup_details = list(agg.values())
                for a in dedup_details:
                    a["start_date"] = start_dt.strftime(fmt)
                    a["end_date"] = end_dt.strftime(fmt)

                total_files_u = len(dedup_details)
                processed_u = sum(1 for x in dedup_details if x.get("success"))
                error_u = total_files_u - processed_u

                results_agg = {
                    "folder": last_folder,
                    "total_files": total_files_u,
                    "processed_files": processed_u,
                    "error_files": error_u,
                    "date_range": {
                        "start": start_dt.strftime("%d/%m/%y %H:%M"),
                        "end": end_dt.strftime("%d/%m/%y %H:%M")
                    },
                    "combined_stats": {
                        "total_rows": int(combined.shape[0]),
                        "total_columns": int(combined.shape[1]),
                        "total_kwh_values": int(pd.notna(combined["kwh"]).sum()) if "kwh" in combined.columns else 0,
                        "total_kvar_values": int(pd.notna(combined["kvarh"]).sum()) if "kvarh" in combined.columns else 0,
                        "resolution": resolution,
                        "rows_before_parse": before_rows,
                        "rows_after_parse": after_parse_rows,
                        "rows_after_filter": after_filter_rows
                    },
                    "file_details": dedup_details,
                    "errors": []
                }
                self.root.after(0, lambda: self.on_analysis_done(True, f"Procesamiento {file_type.upper()} completado", results_agg))
            except Exception as e:
                self.root.after(0, lambda: self.show_error(str(e)))
            finally:
                self.root.after(0, lambda: self.set_busy(False, "Listo"))

        threading.Thread(target=worker, daemon=True).start()

    def append_info(self, text: str):
        self.info_text.configure(state="normal")
        self.info_text.insert(tk.END, text.rstrip() + "\n")
        self.info_text.see(tk.END)
        self.info_text.configure(state="disabled")

    def clear_results(self):
        """Limpia panel de información y deshabilita exportaciones."""
        self.info_text.configure(state="normal")
        self.info_text.delete("1.0", "end")
        self.info_text.configure(state="disabled")
        self.export_excel_btn.configure(state="disabled")
        self.export_csv_btn.configure(state="disabled")
        self.append_info("Panel limpiado.")
        self.last_results = None
        self.csv_processor.combined_df = None

    # --- Utilidades reporte ---
    def populate_companies(self):
        df = getattr(self.csv_processor, "combined_df", None)
        if df is None or df.empty:
            self.company_cb.configure(state="disabled", values=[])
            self.report_btn.configure(state="disabled")
            return
        if "company" in df.columns:
            companies = sorted([str(x) for x in df["company"].dropna().unique().tolist()])
        else:
            companies = ["General"]
            df["company"] = "General"
        self.company_cb.configure(state="readonly", values=companies)
        if not self.company_cb.get():
            self.company_cb.set(companies[0])
        # Al poblar, refleja el multiplo para la empresa actual (si existe)
        self._on_company_selected()
        self.report_btn.configure(state="normal")

    # --- Multiplo por empresa: sincronización UI <-> cache ---
    def _on_company_selected(self, event=None):
        """Cuando el usuario selecciona una empresa, muestra su multiplo guardado
        o el multiplo por defecto si no existe."""
        company = self.company_cb.get()
        if not company:
            return
        val = self.company_multipliers.get(company)
        if val is None:
            val = getattr(self, 'default_multiplier', 80)
        try:
            self.multiplier_sp.set(str(int(val)))
        except Exception:
            self.multiplier_sp.set(str(val))

    def _on_multiplier_change(self):
        """Guarda el multiplo actual para la empresa seleccionada cuando cambia."""
        company = self.company_cb.get()
        if not company:
            return
        try:
            m = float(self.multiplier_sp.get())
        except Exception:
            return
        self.company_multipliers[company] = m

    def _hourly_aggregate(self, df):
        # Asegura timestamp datetime, agrega por hora
        if "timestamp" not in df.columns:
            return df.iloc[0:0].copy()
        if not pd.api.types.is_datetime64_any_dtype(df["timestamp"]):
            df = df.copy()
            df["timestamp"] = pd.to_datetime(df["timestamp"].astype(str), errors="coerce", dayfirst=True)
            df = df[pd.notna(df["timestamp"])]
        df["hour_ts"] = df["timestamp"].dt.floor("H")
        agg_cols = {}
        if "kwh" in df.columns:
            agg_cols["kwh"] = "sum"
        if "kvarh" in df.columns:
            agg_cols["kvarh"] = "sum"
        if not agg_cols:
            return df.iloc[0:0].copy()
        g = df.groupby("hour_ts", as_index=True).agg(agg_cols)
        return g

    def compute_report_table(self, company: str, start_dt: datetime, end_dt: datetime, multiplo: float):
        df = getattr(self.csv_processor, "combined_df", None)
        if df is None or df.empty:
            return pd.DataFrame(), {"kwh": 0.0, "kvarh": 0.0}, {}
        # Filtrar por empresa y rango
        if "company" in df.columns:
            df = df[df["company"].astype(str) == str(company)].copy()
        # Ajustar rango
        if "timestamp" in df.columns:
            if not pd.api.types.is_datetime64_any_dtype(df["timestamp"]):
                df["timestamp"] = pd.to_datetime(df["timestamp"].astype(str), errors="coerce", dayfirst=True)
                df = df[pd.notna(df["timestamp"])]
            df = df[(df["timestamp"] >= start_dt) & (df["timestamp"] <= end_dt)]
        # Agregado por hora
        hourly = self._hourly_aggregate(df)  # index: hour_ts
        # Crear rejilla de días/24h
        rows = []
        cur = datetime(start_dt.year, start_dt.month, start_dt.day)
        end_day = datetime(end_dt.year, end_dt.month, end_dt.day)
        while cur <= end_day:
            for h in range(24):
                ts = cur + timedelta(hours=h)
                kv = hourly.loc[ts] if ts in hourly.index else None
                kwh = float(kv["kwh"]) if kv is not None and "kwh" in hourly.columns else 0.0
                kvarh = float(kv["kvarh"]) if kv is not None and "kvarh" in hourly.columns else 0.0
                # Aplicar multiplo a valores por hora
                kwh_scaled = kwh * multiplo
                kvarh_scaled = kvarh * multiplo
                rows.append({
                    "Fecha": format_es_date(cur.date()),
                    "Hora": h + 1,  # 1..24
                    "Kwh": round(kwh_scaled, 3),
                    "Kvarh": round(kvarh_scaled, 3),
                })
            cur += timedelta(days=1)
        report = pd.DataFrame(rows, columns=["Fecha", "Hora", "Kwh", "Kvarh"])
        totals = {
            # Totales = suma de valores ya multiplicados (NO se vuelve a multiplicar)
            "kwh": float(report["Kwh"].sum()),
            "kvarh": float(report["Kvarh"].sum()),
        }
        meta = {"company": company, "multiplo": multiplo}
        return report, totals, meta

    def generate_report(self):
        df = getattr(self.csv_processor, "combined_df", None)
        if df is None or df.empty:
            messagebox.showinfo("Reporte", "No hay datos para generar reporte.")
            return
        company = self.company_cb.get() or ("General" if "company" not in df.columns else str(df["company"].iloc[0]))
        try:
            m = float(self.multiplier_sp.get())
        except Exception:
            m = 80.0
        # Guardar preferencia por empresa
        self.company_multipliers[company] = m
        # Construir tabla
        sdate = self.start_date.get_date(); edate = self.end_date.get_date()
        sh, sm, eh, em = self._sanitize_time_inputs()
        start_dt = datetime(sdate.year, sdate.month, sdate.day, 0, 0)
        end_dt = datetime(edate.year, edate.month, edate.day, 23, 59)
        report_df, totals, meta = self.compute_report_table(company, start_dt, end_dt, m)
        if report_df.empty:
            messagebox.showinfo("Reporte", "No se generaron filas para el rango seleccionado.")
            return
        self.last_report = {"df": report_df, "totals": totals, "meta": meta}
        self.show_report_window(report_df, totals, meta)

    def show_report_window(self, report_df: pd.DataFrame, totals: dict, meta: dict):
        win = tk.Toplevel(self.root)
        win.title(f"Reporte mensual - {meta.get('company','')}")
        win.geometry("700x680")
        top = ttk.Frame(win, padding=12)
        top.pack(side="top", fill="x")
        ttk.Label(top, text=f"Multiplo → {int(meta.get('multiplo', 0))}", font=("Segoe UI", 12, "bold")).pack(side="left")
        sep = ttk.Frame(top); sep.pack(side="left", expand=True, fill="x")
        ttk.Label(top, text=f"{totals['kwh']:,.3f}", font=("Segoe UI", 14, "bold"), foreground="#1b4f72").pack(side="left", padx=(8, 16))
        ttk.Label(top, text=f"{totals['kvarh']:,.3f}", font=("Segoe UI", 14, "bold"), foreground="#1b4f72").pack(side="left")

        # Tabla
        cols = ("Fecha", "Hora", "Kwh", "Kvarh")
        tree = ttk.Treeview(win, columns=cols, show="headings", height=24)
        for c, w in [("Fecha", 120), ("Hora", 60), ("Kwh", 120), ("Kvarh", 120)]:
            tree.heading(c, text=c)
            tree.column(c, width=w, anchor="e" if c in ("Hora", "Kwh", "Kvarh") else "w")
        # Volcado
        for _, r in report_df.iterrows():
            tree.insert("", "end", values=(r["Fecha"], f"{int(r['Hora'])}", f"{r['Kwh']:.3f}", f"{r['Kvarh']:.3f}"))
        tree.pack(side="top", fill="both", expand=True, padx=12, pady=8)

        # Botones export
        btnf = ttk.Frame(win, padding=12)
        btnf.pack(side="bottom", fill="x")
        ttk.Button(btnf, text="Exportar reporte a Excel", command=self.export_report_excel).pack(side="right")

    def export_report_excel(self):
        if not self.last_report:
            messagebox.showinfo("Exportar", "No hay reporte para exportar.")
            return
        report_df = self.last_report["df"]
        totals = self.last_report["totals"]
        meta = self.last_report["meta"]
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar reporte mensual"
        )
        if not path:
            return
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            wb = Workbook()
            ws = wb.active
            ws.title = "Reporte"

            # Encabezado
            ws["A1"] = "Multiplo →"
            ws["B1"] = int(meta.get("multiplo", 0))
            ws["C1"] = "Kwh"
            ws["D1"] = "Kvarh"
            ws["C2"] = totals["kwh"]
            ws["D2"] = totals["kvarh"]
            ws["C2"].number_format = "#,##0.000"
            ws["D2"].number_format = "#,##0.000"
            ws["A1"].font = Font(bold=True, size=12)
            ws["C2"].font = Font(bold=True, size=14)
            ws["D2"].font = Font(bold=True, size=14)

            # Cabecera de tabla
            headers = ["Fecha", "Hora", "Kwh", "Kvarh"]
            ws.append([])  # fila 3 vacía
            ws.append(headers)  # fila 4
            hdr_fill = PatternFill("solid", fgColor="D9EAF7")
            for col, h in enumerate(headers, start=1):
                cell = ws.cell(row=4, column=col)
                cell.fill = hdr_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Datos
            thin = Side(style="thin", color="999999")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            start_row = 5
            for idx, r in report_df.iterrows():
                row = start_row + idx
                ws.cell(row=row, column=1, value=r["Fecha"])
                ws.cell(row=row, column=2, value=int(r["Hora"]))
                c3 = ws.cell(row=row, column=3, value=float(r["Kwh"]))
                c4 = ws.cell(row=row, column=4, value=float(r["Kvarh"]))
                c3.number_format = "#,##0.000"
                c4.number_format = "#,##0.000"
                for col in range(1, 5):
                    ws.cell(row=row, column=col).border = border
                # Colorear columnas Kwh/Kvarh
                ws.cell(row=row, column=3).fill = PatternFill("solid", fgColor="E9F5FE")
                ws.cell(row=row, column=4).fill = PatternFill("solid", fgColor="E9F5FE")

            # Anchos
            ws.column_dimensions["A"].width = 14
            ws.column_dimensions["B"].width = 8
            ws.column_dimensions["C"].width = 14
            ws.column_dimensions["D"].width = 14

            wb.save(path)
            messagebox.showinfo("Exportar", f"Archivo guardado:\n{path}")
        except Exception as e:
            self.show_error(str(e))

    # ---------- Exportaciones clásicas (Excel/CSV combinados) ----------
    def export_excel(self):
        """Exporta a Excel con una hoja por empresa e incluye Totales y Multiplo al inicio.
        Los totales se calculan como suma(Kwh) y suma(Kvarh) del rango analizado y se multiplican
        por el multiplo de la empresa (si existe) o el actual del spinner como valor por defecto.
        """
        df = getattr(self.csv_processor, "combined_df", None)
        if df is None or df.empty:
            messagebox.showinfo("Exportar", "No hay datos para exportar.")
            return

        # Ruta de guardado
        path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel", "*.xlsx")],
            title="Guardar Excel (multi-hoja)"
        )
        if not path:
            return

        # Solo aplicar el multiplo del spinner a la empresa seleccionada;
        # las demás usan su valor guardado (o 1.0 por defecto)
        selected_company = self.company_cb.get() or None
        try:
            selected_multiplo = float(self.multiplier_sp.get())
        except Exception:
            selected_multiplo = None

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

            # Preparar datos: formatear timestamp a texto día/mes/año HH:MM:SS
            data = df.copy()
            if "timestamp" in data.columns and not data["timestamp"].isna().all():
                if pd.api.types.is_datetime64_any_dtype(data["timestamp"]):
                    data["timestamp"] = data["timestamp"].dt.strftime("%d/%m/%Y %H:%M:%S")
                else:
                    ts = pd.to_datetime(data["timestamp"].astype(str), errors="coerce", dayfirst=True)
                    data["timestamp"] = ts.dt.strftime("%d/%m/%Y %H:%M:%S")

            # Crear libro, estilos y hoja TOTAL (primera)
            wb = Workbook()
            hdr_fill = PatternFill("solid", fgColor="D9EAF7")
            light_fill = PatternFill("solid", fgColor="E9F5FE")
            thin = Side(style="thin", color="999999")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            ws_total = wb.active
            ws_total.title = "total"

            # Cabecera TOTAL
            total_headers = ["No.", "Cliente", "Multiplo", "KWh", "KVARh", "KW"]
            ws_total.append(total_headers)
            for c in range(1, len(total_headers) + 1):
                cell = ws_total.cell(row=1, column=c)
                cell.fill = hdr_fill
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")

            # Preparar compañías y mapa nombre->hoja
            companies = sorted([str(x) for x in data["company"].dropna().unique()]) if "company" in data.columns else ["General"]
            if "company" not in data.columns:
                data["company"] = "General"

            # Helper para nombre de hoja único (cap Excel 31 chars)
            used_titles = set([ws_total.title])
            def unique_title(base: str) -> str:
                t = str(base)[:31]
                if t not in used_titles:
                    used_titles.add(t); return t
                i = 2
                while True:
                    cand = (str(base)[:31-len(str(i))-1] + f" {i}") if len(str(base)) >= 31 else f"{str(base)} {i}"
                    cand = cand[:31]
                    if cand not in used_titles:
                        used_titles.add(cand); return cand
                    i += 1

            sheet_by_company = {}

            # Crear hojas por empresa y recolectar totales + rango para KW
            total_rows = []
            for idx, company in enumerate(companies, start=1):
                 # Resolver multiplo por empresa: usa el guardado; si no existe, el valor por defecto (spinner inicial)
                m = float(self.company_multipliers.get(company, getattr(self, 'default_multiplier', 80)))
                if selected_company and company == selected_company and (selected_multiplo is not None):
                    m = float(selected_multiplo)
                    # Actualizar cache para esta empresa únicamente
                    self.company_multipliers[company] = m
                cdf = data[data["company"].astype(str) == company].copy()

                # Totales multiplicados para TOTAL
                kwh_sum = float(pd.to_numeric(cdf.get("kwh"), errors="coerce").sum()) if "kwh" in cdf.columns else 0.0
                kvar_sum = float(pd.to_numeric(cdf.get("kvarh"), errors="coerce").sum()) if "kvarh" in cdf.columns else 0.0
                kwh_total = kwh_sum * m
                kvar_total = kvar_sum * m

                # Hoja empresa
                sheet_name = unique_title(company)
                ws = wb.create_sheet(title=sheet_name)
                sheet_by_company[company] = sheet_name

                # Encabezado Totales/Multiplo
                ws["A1"] = "Multiplo →"; ws["A1"].font = Font(bold=True, size=12)
                ws["B1"] = int(m)
                # Colocar directamente los totales numéricos en D1 y E1 como solicitaste
                ws["C1"] = "Kwh"; ws["C1"].font = Font(bold=True, size=12)
                ws["D1"] = kwh_total; ws["D1"].number_format = "#,##0.000"; ws["D1"].font = Font(bold=True, size=14)
                ws["E1"] = kvar_total; ws["E1"].number_format = "#,##0.000"; ws["E1"].font = Font(bold=True, size=14)

                # Cabeceras de tabla
                ws.append([])  # fila 3
                cols = []
                for name in ["timestamp", "Hora", "company", "kwh", "kvarh"]:
                    if name in cdf.columns and name not in cols:
                        cols.append(name)
                for name in cdf.columns:
                    if name not in cols:
                        cols.append(name)
                ws.append(cols)  # fila 4
                for col_idx, h in enumerate(cols, start=1):
                    cell = ws.cell(row=4, column=col_idx)
                    cell.fill = hdr_fill
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

                # Datos
                start_row = 5
                for ridx, (_, row) in enumerate(cdf.iterrows(), start=start_row):
                    for cidx, col_name in enumerate(cols, start=1):
                        val = row.get(col_name)
                        cell = ws.cell(row=ridx, column=cidx, value=val if pd.notna(val) else None)
                        cell.border = border
                        if col_name.lower() in ("kwh", "kvarh"):
                            cell.number_format = "#,##0.000"
                            cell.fill = light_fill

                # Anchos
                ws.column_dimensions["A"].width = max(14, min(28, len(str(cols[0])) + 6)) if cols else 14
                ws.column_dimensions["B"].width = 10
                ws.column_dimensions["C"].width = 16
                ws.column_dimensions["D"].width = 16

                # Guardar info para TOTAL (hipervínculo y fórmula KW)
                last_row = start_row + max(len(cdf), 1) - 1
                if len(cdf) <= 0:
                    kw_formula = 0
                else:
                    # Kwh está en columna D por el orden definido
                    esc = sheet_name.replace("'", "''")
                    kw_formula = f"=MAX('{esc}'!$D${start_row}:$D${last_row})"
                total_rows.append((idx, company, m, kwh_total, kvar_total, kw_formula, esc, start_row, last_row))

            # Volcar TOTAL con hipervínculos
            for r_idx, (no, company, m, kwh_t, kvar_t, kw_formula, esc, srow, lrow) in enumerate(total_rows, start=2):
                ws_total.cell(row=r_idx, column=1, value=no)
                c_name = ws_total.cell(row=r_idx, column=2, value=company)
                sheet_name = sheet_by_company[company]
                esc = sheet_name.replace("'", "''")
                # Hipervínculo con nombre de hoja entre comillas simples
                c_name.hyperlink = f"#'{esc}'!A1"
                c_name.font = Font(color="0563C1", underline="single")
                ws_total.cell(row=r_idx, column=3, value=int(m))
                # KWh/KVARh en TOTAL con respaldo: si D1/E1 no son números, sumar columna de datos
                c4 = ws_total.cell(row=r_idx, column=4)
                c4.value = f"=IF(ISNUMBER('{esc}'!$D$1), '{esc}'!$D$1, SUM('{esc}'!$D${srow}:'{esc}'!$D${lrow}))"; c4.number_format = "#,##0.000"
                c5 = ws_total.cell(row=r_idx, column=5)
                c5.value = f"=IF(ISNUMBER('{esc}'!$E$1), MAX('{esc}'!$E$1), SUM('{esc}'!$E${srow}:'{esc}'!$E${lrow}))"; c5.number_format = "#,##0.000"
                c6 = ws_total.cell(row=r_idx, column=6)
                if isinstance(kw_formula, str):
                    c6.value = kw_formula
                else:
                    c6.value = 0
                c6.number_format = "#,##0.000"

            # Anchos TOTAL
            ws_total.column_dimensions["A"].width = 6
            ws_total.column_dimensions["B"].width = 34
            ws_total.column_dimensions["C"].width = 10
            ws_total.column_dimensions["D"].width = 16
            ws_total.column_dimensions["E"].width = 16
            ws_total.column_dimensions["F"].width = 12

            wb.save(path)
            messagebox.showinfo("Exportar", f"Excel exportado: {path}")
        except Exception as e:
            self.show_error(str(e))

    def export_csv(self):
        df = getattr(self.csv_processor, "combined_df", None)
        if df is None or df.empty:
            messagebox.showinfo("Exportar", "No hay datos para exportar.")
            return
        path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV", "*.csv")],
            title="Guardar CSV combinado"
        )
        if not path:
            return
        ok, msg = self.csv_processor.export_combined_csv(path)
        if ok:
            messagebox.showinfo("Exportar", msg)
        else:
            messagebox.showerror("Exportar", msg)

    # ...existing code on_analysis_done...
    def on_analysis_done(self, ok: bool, msg: str, results: dict):
        self.append_info(msg)
        if ok and getattr(self.csv_processor, "combined_df", None) is not None:
            self.last_results = results
            if hasattr(self, "export_excel_btn"):
                self.export_excel_btn.configure(state="normal")
            if hasattr(self, "export_csv_btn"):
                self.export_csv_btn.configure(state="normal")
            # Habilitar selección de empresa y reporte
            self.populate_companies()
            cs = results.get("combined_stats", {})
            self.append_info(f"Filas: {cs.get('total_rows', 0)}  Columnas: {cs.get('total_columns', 0)}  Resolución: {cs.get('resolution', '')}")
        else:
            self.append_info("Sin resultados para exportar.")
            self.company_cb.configure(state="disabled")
            self.report_btn.configure(state="disabled")
